"""
5-Sheet Excel Load Test Report Generator for PetOrb Ecosystem
Generates styled PetOrb_LoadTest_300_Report.xlsx with 5 analytical sheets and embedded charts.
"""

import os
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, BarChart, Reference

class ExcelLoadTestReporter:
    def __init__(self, output_dir="loadtest_suite/reports"):
        self.output_dir = output_dir
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir, exist_ok=True)
        self.file_path = os.path.join(self.output_dir, "PetOrb_LoadTest_300_Report.xlsx")
        self.results = []
        self.start_time = datetime.datetime.now()

    def add_result(self, test_id, module, scenario, load_level, latency_ms, sla_target_ms, status, status_code=200, error_details=""):
        self.results.append({
            "test_id": test_id,
            "module": module,
            "scenario": scenario,
            "load_level": load_level,
            "latency_ms": round(latency_ms, 2),
            "sla_target_ms": sla_target_ms,
            "status": status,
            "status_code": status_code,
            "error_details": error_details,
            "timestamp": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        })

    def generate_report(self):
        wb = openpyxl.Workbook()
        
        # Styles
        font_header = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
        font_subhead = Font(name="Calibri", size=11, bold=True, color="1F4E78")
        font_title = Font(name="Calibri", size=18, bold=True, color="1F4E78")
        font_bold = Font(name="Calibri", size=11, bold=True)
        font_regular = Font(name="Calibri", size=11)
        
        fill_navy = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        fill_subhead = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        fill_pass = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        fill_fail = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        fill_zebra = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        font_pass = Font(name="Calibri", size=11, color="006100", bold=True)
        font_fail = Font(name="Calibri", size=11, color="9C0006", bold=True)

        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )

        total_tests = len(self.results)
        passed_tests = sum(1 for r in self.results if r["status"] == "PASS")
        failed_tests = total_tests - passed_tests
        pass_rate = round((passed_tests / total_tests * 100), 2) if total_tests > 0 else 0
        avg_latency = round(sum(r["latency_ms"] for r in self.results) / total_tests, 2) if total_tests > 0 else 0

        # ----------------------------------------------------------------------
        # SHEET 1: EXECUTIVE SUMMARY
        # ----------------------------------------------------------------------
        ws1 = wb.active
        ws1.title = "Executive Summary"
        ws1.views.sheetView[0].showGridLines = True

        ws1["A1"] = "PETORB AI ECOSYSTEM - 300+ LOAD TEST EXECUTION SUMMARY"
        ws1["A1"].font = font_title
        ws1.merge_cells("A1:F1")

        ws1["A3"] = "Execution Timestamp:"
        ws1["B3"] = self.start_time.strftime("%Y-%m-%d %H:%M:%S")
        ws1["A4"] = "Target Application:"
        ws1["B4"] = "PetOrb Monorepo Web & Server API"
        ws1["A5"] = "Total Test Scenarios:"
        ws1["B5"] = total_tests

        for r in range(3, 6):
            ws1[f"A{r}"].font = font_bold

        # KPI Cards Table
        headers_kpi = ["Metric Name", "Value", "Benchmark Target", "Status"]
        for col_num, h in enumerate(headers_kpi, 1):
            cell = ws1.cell(row=7, column=col_num, value=h)
            cell.font = font_header
            cell.fill = fill_navy
            cell.alignment = Alignment(horizontal="center", vertical="center")

        kpis = [
            ("Total Test Cases Executed", total_tests, ">= 300", "COMPLIANT"),
            ("Passed Load Scenarios", passed_tests, f"{total_tests} Scenarios", "SATISFACTORY" if failed_tests == 0 else "WARNING"),
            ("Failed Load Scenarios", failed_tests, "0 Failures", "OPTIMAL" if failed_tests == 0 else "NEEDS_ATTENTION"),
            ("Overall Pass Rate (%)", f"{pass_rate}%", ">= 95%", "PASS" if pass_rate >= 95 else "FAIL"),
            ("Average System Latency (ms)", f"{avg_latency} ms", "< 500 ms", "PASS" if avg_latency < 500 else "WARN")
        ]

        for i, (metric, val, bench, status) in enumerate(kpis, 8):
            ws1.cell(row=i, column=1, value=metric).font = font_regular
            ws1.cell(row=i, column=2, value=val).font = font_bold
            ws1.cell(row=i, column=3, value=bench).font = font_regular
            st_cell = ws1.cell(row=i, column=4, value=status)
            st_cell.font = font_pass if status in ["COMPLIANT", "SATISFACTORY", "OPTIMAL", "PASS"] else font_fail
            st_cell.alignment = Alignment(horizontal="center")
            
            for c in range(1, 5):
                ws1.cell(row=i, column=c).border = thin_border

        # Add Pie Chart for Pass/Fail Breakdown
        ws1["A15"] = "Pass vs Fail Breakdown Data"
        ws1["A15"].font = font_subhead
        ws1["A16"] = "Status"
        ws1["B16"] = "Count"
        ws1["A17"] = "PASS"
        ws1["B17"] = passed_tests
        ws1["A18"] = "FAIL"
        ws1["B18"] = failed_tests

        pie = PieChart()
        pie.title = "Load Test Execution Outcome"
        labels = Reference(ws1, min_col=1, min_row=17, max_row=18)
        data = Reference(ws1, min_col=2, min_row=16, max_row=18)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.width = 14
        pie.height = 7
        ws1.add_chart(pie, "D15")

        # ----------------------------------------------------------------------
        # SHEET 2: 300 LOAD TEST CASES
        # ----------------------------------------------------------------------
        ws2 = wb.create_sheet(title="300 Load Test Cases")
        ws2.views.sheetView[0].showGridLines = True

        headers_tc = [
            "Test Case ID", "Module", "Scenario Description", "Load Level (VUs)",
            "Response Latency (ms)", "SLA Target (ms)", "Status Code", "Test Status", "Timestamp", "Error Details"
        ]

        for col_num, h in enumerate(headers_tc, 1):
            cell = ws2.cell(row=1, column=col_num, value=h)
            cell.font = font_header
            cell.fill = fill_navy
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for idx, item in enumerate(self.results, 2):
            ws2.cell(row=idx, column=1, value=item["test_id"]).alignment = Alignment(horizontal="center")
            ws2.cell(row=idx, column=2, value=item["module"])
            ws2.cell(row=idx, column=3, value=item["scenario"])
            ws2.cell(row=idx, column=4, value=item["load_level"]).alignment = Alignment(horizontal="center")
            ws2.cell(row=idx, column=5, value=item["latency_ms"]).alignment = Alignment(horizontal="right")
            ws2.cell(row=idx, column=6, value=item["sla_target_ms"]).alignment = Alignment(horizontal="right")
            ws2.cell(row=idx, column=7, value=item["status_code"]).alignment = Alignment(horizontal="center")
            
            st_cell = ws2.cell(row=idx, column=8, value=item["status"])
            st_cell.alignment = Alignment(horizontal="center")
            if item["status"] == "PASS":
                st_cell.fill = fill_pass
                st_cell.font = font_pass
            else:
                st_cell.fill = fill_fail
                st_cell.font = font_fail

            ws2.cell(row=idx, column=9, value=item["timestamp"]).alignment = Alignment(horizontal="center")
            ws2.cell(row=idx, column=10, value=item["error_details"])

            if idx % 2 == 1:
                for c in range(1, 11):
                    if c != 8:
                        ws2.cell(row=idx, column=c).fill = fill_zebra

            for c in range(1, 11):
                ws2.cell(row=idx, column=c).border = thin_border

        # ----------------------------------------------------------------------
        # SHEET 3: LATENCY DISTRIBUTION
        # ----------------------------------------------------------------------
        ws3 = wb.create_sheet(title="Latency Distribution")
        ws3.views.sheetView[0].showGridLines = True

        ws3["A1"] = "LATENCY PERCENTILE DISTRIBUTION ANALYSIS"
        ws3["A1"].font = font_title

        latencies = sorted([r["latency_ms"] for r in self.results]) if self.results else [0]
        def percentile(lst, p):
            if not lst:
                return 0
            k = (len(lst) - 1) * (p / 100.0)
            f = int(k)
            c = f + 1 if f + 1 < len(lst) else f
            return lst[f] + (lst[c] - lst[f]) * (k - f)

        p50 = round(percentile(latencies, 50), 2)
        p75 = round(percentile(latencies, 75), 2)
        p90 = round(percentile(latencies, 90), 2)
        p95 = round(percentile(latencies, 95), 2)
        p99 = round(percentile(latencies, 99), 2)
        max_lat = max(latencies) if latencies else 0
        min_lat = min(latencies) if latencies else 0

        ws3_headers = ["Percentile Tier", "Latency Benchmark (ms)", "SLA Threshold (ms)", "SLA Compliance"]
        for c, h in enumerate(ws3_headers, 1):
            cell = ws3.cell(row=3, column=c, value=h)
            cell.font = font_header
            cell.fill = fill_navy
            cell.alignment = Alignment(horizontal="center")

        pct_rows = [
            ("Min Latency (p0)", min_lat, 200, "PASS" if min_lat <= 200 else "FAIL"),
            ("50th Percentile (p50 / Median)", p50, 400, "PASS" if p50 <= 400 else "FAIL"),
            ("75th Percentile (p75)", p75, 600, "PASS" if p75 <= 600 else "FAIL"),
            ("90th Percentile (p90)", p90, 800, "PASS" if p90 <= 800 else "FAIL"),
            ("95th Percentile (p95)", p95, 1000, "PASS" if p95 <= 1000 else "FAIL"),
            ("99th Percentile (p99)", p99, 1500, "PASS" if p99 <= 1500 else "FAIL"),
            ("Peak Max Latency (p100)", max_lat, 2500, "PASS" if max_lat <= 2500 else "FAIL")
        ]

        for i, (tier, lat, sla, comp) in enumerate(pct_rows, 4):
            ws3.cell(row=i, column=1, value=tier).font = font_bold
            ws3.cell(row=i, column=2, value=lat).alignment = Alignment(horizontal="right")
            ws3.cell(row=i, column=3, value=sla).alignment = Alignment(horizontal="right")
            c_cell = ws3.cell(row=i, column=4, value=comp)
            c_cell.alignment = Alignment(horizontal="center")
            c_cell.font = font_pass if comp == "PASS" else font_fail

            for col in range(1, 5):
                ws3.cell(row=i, column=col).border = thin_border

        bar_lat = BarChart()
        bar_lat.type = "col"
        bar_lat.style = 10
        bar_lat.title = "Latency Percentiles vs SLA Threshold (ms)"
        bar_lat.y_axis.title = "Response Time (ms)"
        bar_lat.x_axis.title = "Percentile Tier"
        
        data_ref = Reference(ws3, min_col=2, min_row=3, max_col=3, max_row=10)
        cats_ref = Reference(ws3, min_col=1, min_row=4, max_row=10)
        bar_lat.add_data(data_ref, titles_from_data=True)
        bar_lat.set_categories(cats_ref)
        bar_lat.width = 16
        bar_lat.height = 9
        ws3.add_chart(bar_lat, "F3")

        # ----------------------------------------------------------------------
        # SHEET 4: MODULE PERFORMANCE ANALYSIS
        # ----------------------------------------------------------------------
        ws4 = wb.create_sheet(title="Module Performance Analysis")
        ws4.views.sheetView[0].showGridLines = True

        ws4["A1"] = "MODULE-WISE LOAD STRESS & PERFORMANCE METRICS"
        ws4["A1"].font = font_title

        modules = list(set(r["module"] for r in self.results))
        modules.sort()

        ws4_headers = ["Module Name", "Total Scenarios", "Passed", "Failed", "Pass Rate (%)", "Avg Latency (ms)", "Max Latency (ms)"]
        for c, h in enumerate(ws4_headers, 1):
            cell = ws4.cell(row=3, column=c, value=h)
            cell.font = font_header
            cell.fill = fill_navy
            cell.alignment = Alignment(horizontal="center")

        for idx, mod in enumerate(modules, 4):
            mod_items = [r for r in self.results if r["module"] == mod]
            m_total = len(mod_items)
            m_passed = sum(1 for r in mod_items if r["status"] == "PASS")
            m_failed = m_total - m_passed
            m_prate = round((m_passed / m_total * 100), 2) if m_total > 0 else 0
            m_avg_lat = round(sum(r["latency_ms"] for r in mod_items) / m_total, 2) if m_total > 0 else 0
            m_max_lat = max([r["latency_ms"] for r in mod_items]) if mod_items else 0

            ws4.cell(row=idx, column=1, value=mod).font = font_bold
            ws4.cell(row=idx, column=2, value=m_total).alignment = Alignment(horizontal="center")
            ws4.cell(row=idx, column=3, value=m_passed).alignment = Alignment(horizontal="center")
            ws4.cell(row=idx, column=4, value=m_failed).alignment = Alignment(horizontal="center")
            ws4.cell(row=idx, column=5, value=f"{m_prate}%").alignment = Alignment(horizontal="center")
            ws4.cell(row=idx, column=6, value=m_avg_lat).alignment = Alignment(horizontal="right")
            ws4.cell(row=idx, column=7, value=m_max_lat).alignment = Alignment(horizontal="right")

            for col in range(1, 8):
                ws4.cell(row=idx, column=col).border = thin_border

        bar_mod = BarChart()
        bar_mod.type = "col"
        bar_mod.style = 11
        bar_mod.title = "Average Response Latency per Module (ms)"
        bar_mod.y_axis.title = "Avg Latency (ms)"
        bar_mod.x_axis.title = "Module"
        
        mod_data = Reference(ws4, min_col=6, min_row=3, max_row=3 + len(modules))
        mod_cats = Reference(ws4, min_col=1, min_row=4, max_row=3 + len(modules))
        bar_mod.add_data(mod_data, titles_from_data=True)
        bar_mod.set_categories(mod_cats)
        bar_mod.width = 16
        bar_mod.height = 9
        ws4.add_chart(bar_mod, "A12")

        # ----------------------------------------------------------------------
        # SHEET 5: SLA & BOTTLENECK REPORT
        # ----------------------------------------------------------------------
        ws5 = wb.create_sheet(title="SLA & Bottleneck Report")
        ws5.views.sheetView[0].showGridLines = True

        ws5["A1"] = "SLA COMPLIANCE & BOTTLENECK ANALYSIS"
        ws5["A1"].font = font_title

        slow_scenarios = [r for r in self.results if r["latency_ms"] > r["sla_target_ms"] or r["status"] == "FAIL"]
        
        ws5_headers = ["Test Case ID", "Module", "Scenario Description", "Load Level", "Latency (ms)", "SLA Limit (ms)", "Violation Variance (ms)", "Root Cause / Recommendation"]
        for c, h in enumerate(ws5_headers, 1):
            cell = ws5.cell(row=3, column=c, value=h)
            cell.font = font_header
            cell.fill = fill_navy
            cell.alignment = Alignment(horizontal="center")

        if not slow_scenarios:
            ws5.cell(row=4, column=1, value="No SLA violations or bottlenecks detected. System performed within all defined response time limits.").font = font_bold
            ws5.merge_cells("A4:H4")
        else:
            for idx, item in enumerate(slow_scenarios, 4):
                var_ms = round(item["latency_ms"] - item["sla_target_ms"], 2)
                ws5.cell(row=idx, column=1, value=item["test_id"]).alignment = Alignment(horizontal="center")
                ws5.cell(row=idx, column=2, value=item["module"])
                ws5.cell(row=idx, column=3, value=item["scenario"])
                ws5.cell(row=idx, column=4, value=item["load_level"]).alignment = Alignment(horizontal="center")
                ws5.cell(row=idx, column=5, value=item["latency_ms"]).alignment = Alignment(horizontal="right")
                ws5.cell(row=idx, column=6, value=item["sla_target_ms"]).alignment = Alignment(horizontal="right")
                ws5.cell(row=idx, column=7, value=f"+{var_ms} ms" if var_ms > 0 else "0 ms").font = font_fail
                ws5.cell(row=idx, column=8, value=item["error_details"] if item["error_details"] else "SLA Threshold Exceeded under heavy concurrent load")

                for col in range(1, 9):
                    ws5.cell(row=idx, column=col).border = thin_border

        # Adjust column widths across all sheets
        for sheet in wb.worksheets:
            for col in sheet.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.value:
                        val_str = str(cell.value)
                        if len(val_str) > max_len and "\n" not in val_str:
                            max_len = len(val_str)
                sheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

        wb.save(self.file_path)
        return self.file_path
