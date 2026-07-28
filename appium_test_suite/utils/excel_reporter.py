"""
Excel Analysis Report Generator using openpyxl for PetOrb Appium Automation Test Suite
Generates styled Appium_Test_Execution_Report.xlsx with Dashboard metrics & 300 detailed test results.
"""

import os
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class ExcelReporter:
    def __init__(self, output_dir="reports"):
        self.output_dir = output_dir
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)
        self.file_path = os.path.join(self.output_dir, "Appium_Test_Execution_Report.xlsx")
        self.test_results = []

    def add_result(self, test_id, module, title, category, status, duration_ms, details=""):
        self.test_results.append({
            "test_id": test_id,
            "module": module,
            "title": title,
            "category": category,
            "status": status,
            "duration_ms": duration_ms,
            "timestamp": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "details": details
        })

    def generate_report(self):
        wb = openpyxl.Workbook()
        
        # -------------------------------------------------------------
        # TAB 1: Summary Dashboard
        # -------------------------------------------------------------
        ws_dash = wb.active
        ws_dash.title = "Executive Summary"
        ws_dash.views.sheetView[0].showGridLines = True

        total_tests = len(self.test_results)
        passed_tests = sum(1 for r in self.test_results if r['status'] == 'PASS')
        failed_tests = sum(1 for r in self.test_results if r['status'] == 'FAIL')
        skipped_tests = sum(1 for r in self.test_results if r['status'] == 'SKIP')
        pass_rate = (passed_tests / total_tests * 100) if total_tests > 0 else 0

        # Styles
        font_header = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
        font_sub = Font(name="Calibri", size=11, bold=True, color="1F4E78")
        font_bold = Font(name="Calibri", size=11, bold=True)
        fill_navy = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        fill_header = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        fill_green = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        fill_red = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )

        # Title Block
        ws_dash.merge_cells("A1:E2")
        title_cell = ws_dash["A1"]
        title_cell.value = "PetOrb Appium E2E Automation - Executive Report"
        title_cell.font = font_header
        title_cell.fill = fill_navy
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Summary KPIs
        kpis = [
            ("Execution Target", "PetOrb Cross-Platform (Flutter / Web / Android)"),
            ("Test Date", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("Total Test Cases", total_tests),
            ("Passed", passed_tests),
            ("Failed", failed_tests),
            ("Skipped", skipped_tests),
            ("Pass Rate (%)", f"{pass_rate:.1f}%")
        ]

        ws_dash.cell(row=4, column=1, value="Metric").font = font_bold
        ws_dash.cell(row=4, column=2, value="Value").font = font_bold
        ws_dash.cell(row=4, column=1).fill = fill_header
        ws_dash.cell(row=4, column=2).fill = fill_header

        for idx, (k, v) in enumerate(kpis, start=5):
            c1 = ws_dash.cell(row=idx, column=1, value=k)
            c2 = ws_dash.cell(row=idx, column=2, value=v)
            c1.border = thin_border
            c2.border = thin_border
            c1.font = font_bold
            if k == "Passed":
                c2.fill = fill_green
            elif k == "Failed" and failed_tests > 0:
                c2.fill = fill_red

        # Module Breakdown Table
        ws_dash.cell(row=14, column=1, value="Module Name").font = font_bold
        ws_dash.cell(row=14, column=2, value="Total Tests").font = font_bold
        ws_dash.cell(row=14, column=3, value="Passed").font = font_bold
        ws_dash.cell(row=14, column=4, value="Failed").font = font_bold
        ws_dash.cell(row=14, column=5, value="Pass Rate (%)").font = font_bold

        for col in range(1, 6):
            ws_dash.cell(row=14, column=col).fill = fill_header

        modules = list(set(r['module'] for r in self.test_results))
        for row_idx, mod in enumerate(modules, start=15):
            mod_tests = [r for r in self.test_results if r['module'] == mod]
            tot = len(mod_tests)
            pas = sum(1 for r in mod_tests if r['status'] == 'PASS')
            fai = sum(1 for r in mod_tests if r['status'] == 'FAIL')
            rate = (pas / tot * 100) if tot > 0 else 0

            ws_dash.cell(row=row_idx, column=1, value=mod).border = thin_border
            ws_dash.cell(row=row_idx, column=2, value=tot).border = thin_border
            ws_dash.cell(row=row_idx, column=3, value=pas).border = thin_border
            ws_dash.cell(row=row_idx, column=4, value=fai).border = thin_border
            ws_dash.cell(row=row_idx, column=5, value=f"{rate:.1f}%").border = thin_border

        # Auto-fit columns in Dashboard
        for col in ws_dash.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws_dash.column_dimensions[col_letter].width = max(max_len + 4, 18)

        # -------------------------------------------------------------
        # TAB 2: Detailed 300 Test Case Results
        # -------------------------------------------------------------
        ws_details = wb.create_sheet(title="300 Test Cases Breakdown")
        ws_details.views.sheetView[0].showGridLines = True

        headers = [
            "Test ID", "Module", "Test Title", "Category", 
            "Status", "Duration (ms)", "Timestamp", "Execution Log / Details"
        ]

        ws_details.append(headers)
        for col_num, header in enumerate(headers, 1):
            cell = ws_details.cell(row=1, column=col_num)
            cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            cell.fill = fill_navy
            cell.alignment = Alignment(horizontal="center", vertical="center")

        fill_pass = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") # Soft Green
        fill_fail = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # Soft Red
        fill_skip = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid") # Soft Yellow

        font_pass = Font(color="006100", bold=True)
        font_fail = Font(color="9C0006", bold=True)
        font_skip = Font(color="9C6500", bold=True)

        for row_idx, res in enumerate(self.test_results, start=2):
            ws_details.append([
                res['test_id'],
                res['module'],
                res['title'],
                res['category'],
                res['status'],
                res['duration_ms'],
                res['timestamp'],
                res['details']
            ])

            status_cell = ws_details.cell(row=row_idx, column=5)
            if res['status'] == 'PASS':
                status_cell.fill = fill_pass
                status_cell.font = font_pass
            elif res['status'] == 'FAIL':
                status_cell.fill = fill_fail
                status_cell.font = font_fail
            else:
                status_cell.fill = fill_skip
                status_cell.font = font_skip

            for c in range(1, 9):
                ws_details.cell(row=row_idx, column=c).border = thin_border

        for col in ws_details.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws_details.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 60)

        wb.save(self.file_path)
        print(f"[ExcelReporter] Successfully generated report: {self.file_path}")
        return self.file_path
