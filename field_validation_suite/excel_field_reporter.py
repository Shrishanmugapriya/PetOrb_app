import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference

class ExcelFieldValidationReporter:
    def __init__(self, output_dir=None):
        if output_dir is None:
            output_dir = os.path.join(os.path.dirname(__file__), "reports")
        os.makedirs(output_dir, exist_ok=True)
        self.output_dir = output_dir
        self.results = []

    def add_result(self, test_id, module, field_name, input_value, expected_rule, status, latency_ms=0.0):
        self.results.append({
            "test_id": test_id,
            "module": module,
            "field_name": field_name,
            "input_value": str(input_value),
            "expected_rule": expected_rule,
            "status": status,
            "latency_ms": round(latency_ms, 2)
        })

    def generate_report(self):
        wb = openpyxl.Workbook()
        
        # Styles
        header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        title_font = Font(name="Calibri", size=16, bold=True, color="1F497D")
        sub_font = Font(name="Calibri", size=11, italic=True, color="595959")
        pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        pass_font = Font(name="Calibri", size=11, bold=True, color="006100")
        fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        fail_font = Font(name="Calibri", size=11, bold=True, color="9C0006")
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )

        # ---------------------------------------------------------
        # SHEET 1: Executive Dashboard
        # ---------------------------------------------------------
        ws_dash = wb.active
        ws_dash.title = "Executive Summary"
        ws_dash.views.sheetView[0].showGridLines = True

        ws_dash["A1"] = "PetOrb Field Validation Automation Summary"
        ws_dash["A1"].font = title_font
        ws_dash["A2"] = "300 Unique Input Schema & Field Validation Test Results"
        ws_dash["A2"].font = sub_font

        total_tests = len(self.results)
        passed = sum(1 for r in self.results if r["status"] == "PASS")
        failed = total_tests - passed
        pass_rate = round((passed / total_tests * 100), 2) if total_tests > 0 else 0

        dash_headers = ["Metric", "Value"]
        ws_dash.append([])
        ws_dash.append(dash_headers)
        dash_row_start = ws_dash.max_row
        
        for col_num in range(1, 3):
            cell = ws_dash.cell(row=dash_row_start, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        summary_metrics = [
            ("Total Test Cases", total_tests),
            ("Passed Validations", passed),
            ("Failed Validations", failed),
            ("Pass Percentage", f"{pass_rate}%")
        ]

        for m_name, m_val in summary_metrics:
            ws_dash.append([m_name, m_val])
            r_idx = ws_dash.max_row
            ws_dash.cell(row=r_idx, column=1).font = Font(name="Calibri", size=11, bold=True)
            ws_dash.cell(row=r_idx, column=2).alignment = Alignment(horizontal="center")
            for c in range(1, 3):
                ws_dash.cell(row=r_idx, column=c).border = thin_border

        # Pie Chart
        pie = PieChart()
        pie.title = "Field Validation Execution Status"
        labels = Reference(ws_dash, min_col=1, min_row=5, max_row=6)
        data = Reference(ws_dash, min_col=2, min_row=4, max_row=6)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.width = 14
        pie.height = 7.5
        ws_dash.add_chart(pie, "D4")

        # ---------------------------------------------------------
        # SHEET 2: Detailed Results
        # ---------------------------------------------------------
        ws_det = wb.create_sheet(title="Field Validation Details")
        ws_det.views.sheetView[0].showGridLines = True

        headers = ["Test ID", "Module", "Field Name", "Input Value", "Expected Validation Rule", "Status", "Latency (ms)"]
        ws_det.append(headers)
        for col_num in range(1, len(headers) + 1):
            cell = ws_det.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for r in self.results:
            ws_det.append([
                r["test_id"],
                r["module"],
                r["field_name"],
                r["input_value"],
                r["expected_rule"],
                r["status"],
                r["latency_ms"]
            ])
            row_idx = ws_det.max_row
            status_cell = ws_det.cell(row=row_idx, column=6)
            if r["status"] == "PASS":
                status_cell.fill = pass_fill
                status_cell.font = pass_font
            else:
                status_cell.fill = fail_fill
                status_cell.font = fail_font
            status_cell.alignment = Alignment(horizontal="center")

            for c in range(1, len(headers) + 1):
                ws_det.cell(row=row_idx, column=c).border = thin_border

        # Auto column widths
        for ws in [ws_dash, ws_det]:
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        file_path = os.path.join(self.output_dir, "Field_Validation_Test_Report.xlsx")
        wb.save(file_path)
        return file_path
